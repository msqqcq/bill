from tinydb import TinyDB, Query
from flask import Flask, make_response, jsonify, request
from datetime import datetime
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import logging
import json
import os

app = Flask(__name__)
CORS(app, methods=["GET", "POST"])

current_dir = os.path.dirname(__file__)
db_dir = os.path.join(current_dir, 'db/')
os.makedirs(db_dir, exist_ok=True)
dpm_db = db_dir + "department.json"
dpm_table = "Department"
product_db = db_dir + "product.json"
product_table = "Product"

sale_table = "Sale"

log = logging.getLogger('flask_app')
log.setLevel(logging.DEBUG)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)
log.addHandler(console_handler)

row_count = 0


def init():
    with open(product_db, 'r') as file:
        data = eval(json.dumps(json.load(file)))

    for item in data["Product"]:
        if 'type' not in data["Product"][item]:
            data["Product"][item]['type'] = 'app'

    with open(product_db, 'w') as file:
        json.dump(data, file)


def file(filename, data):
    response = make_response(data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename={}'.format(filename.encode('utf-8'))
    return response


def ok(msg='', data=None):
    response_data = {
        'message': msg,
        'data': data
    }
    response = make_response(jsonify(response_data))
    response.headers['Content-Type'] = 'application/json'
    response.status_code = 200
    return response


def error(code=500, msg=''):
    response_data = {
        'message': msg
    }
    response = make_response(jsonify(response_data))
    response.headers['Content-Type'] = 'application/json'
    response.status_code = code
    return response


@app.route('/', methods=['GET'])
def holle():
    return datetime.now().strftime("%Y%m%d")


@app.route('/add/dpm', methods=['POST'])
def add_dpm():
    try:
        names = request.json
        if names is None or len(names) == 0:
            return error(400, "请填写数据")

        dups = check_duplicate_dpm(names)
        if len(dups) > 0:
            return error(409, json.dumps(dups) + "已经存在")

        add_departments(names)
        return ok()
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/rm/dpm/<name>', methods=['POST'])
def rm_dpm(name):
    try:
        if is_name_null(name):
            return error(400, "请填写数据")

        # dups = check_duplicate_dpm(names)
        # if len(dups) == 0:
        #     return error(409, json.dumps(dups) + "不存在")

        del_department(name)
        return ok('[{}] 已删除'.format(name))
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/list/dpm', methods=['GET'])
def list_dpm():
    try:
        return ok(data=get_department())
    except Exception as e:
        return error(500, "服务器出错")


def check_duplicate_dpm(names):
    dpms = get_department()
    return [name for name in names if name in dpms]


@app.route('/add/product', methods=['POST'])
def add_prod():
    try:
        p = request.json
        if p is None or is_name_null(p['name']) or p['price'] is None:
            return error(400, "请填写数据")

        for pp in get_product():
            if pp['name'] == p['name']:
                return error(409, "[{}] 已存在".format(p['name']))

        add_products([p])
        return ok()
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/update/product', methods=['POST'])
def update_prod():
    try:
        p = request.json
        if p is None or is_name_null(p['name']) or p['price'] is None:
            return error(400, "请填写数据")
        else:
            for pp in get_product():
                if pp['name'] == p['name']:
                    update_product(p)
                    return ok()

        return error(409, "[{}] 不存在".format(p['name']))
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/rm/product/<name>', methods=['POST'])
def rm_prod(name):
    try:
        if is_name_null(name):
            return error(400, "请填写数据")
        del_product(name)
        return ok('[{}] 已删除'.format(name))
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/list/product', methods=['GET'])
def list_pd():
    try:
        return ok(data=get_product())
    except Exception as e:
        return error(500, "服务器出错")


@app.route('/add/sale/<date>', methods=['POST'])
def add_sale(date):
    try:
        sales = request.json
        if date == '' or sales is None or len(sales) == 0:
            return error(400, "请填写数据")

        sales = [s for s in sales if s['num'] > 0]
        clear_sales(date, sales[0]['dpm'])
        add_sales(date, sales)
        return ok()
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/rm/sale/<date>', methods=['POST'])
def rm_sale(date):
    try:
        dpms = request.json
        if date == '':
            return error(400, "请填写日期")

        if dpms is None or len(dpms) == 0:
            dpms = ['']
        for dpm in dpms:
            clear_sales(date, dpm)
        return ok()
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/list/sale/<date>', methods=['GET'])
def list_sale(date):
    try:
        if date == '':
            return error(400, "请填写日期")
        return ok(data=get_sales(date))
    except Exception as e:
        return error(500, "服务器出错")


@app.route('/check/sale/<date>/<dpm>', methods=['GET'])
def check_sale(date, dpm):
    try:
        if date == '':
            return error(400, "请填写日期")

        data = generate_table(date, dpm)
        return ok(data=data)
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


@app.route('/download/<date>/<dpm>', methods=['GET'])
def download(date, dpm):
    try:
        if date == '':
            return error(400, "请填写日期")
        bytes = generate_excel(dpm, generate_table(date, dpm), date)
        if is_name_null(dpm):
            filename = '总表_{}.xlsx'.format(dpm, date)
        else:
            filename = '{}_{}.xlsx'.format(dpm, date)
        return file(filename, bytes)
    except Exception as e:
        log.error(e)
        return error(500, "服务器出错")


def generate_excel(dpm, data, date):
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bule_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    # 创建工作簿和表格
    wb = Workbook()
    ws = wb.active

    cells = get_excel_columns()

    for i in range(0, len(data)):
        row_data = data[i]
        for j in range(0, len(row_data)):
            idx = cells[j] + str(i + 2)
            ws[idx] = row_data[j]
            # 颜色
            if not is_name_null(dpm):
                if row_data[1] == '收入':
                    ws[idx].fill = green_fill
                elif row_data[1] == '总计':
                    ws[idx].fill = bule_fill
            else:
                if j == len(row_data) - 1:
                    ws[idx].fill = green_fill
                elif (i < len(data) - 2 and j == len(row_data) - 3) or (row_data[0] == '总计' and j < len(row_data) - 3):
                    ws[idx].fill = bule_fill

    # 适配所有列宽
    for column in ws.columns:
        max_length = 5
        column = [cell for cell in column]
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = 2 * max_length
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # 添加 title
    if not is_name_null(dpm):
        ws['A1'] = '{} {}'.format(dpm, date)
    else:
        ws['A1'] = date + ' 收入'
    ws.merge_cells('A1:{}1'.format(cells[len(data[0]) - 1]))
    merge_cell = ws['A1']
    merge_cell.alignment = Alignment(horizontal='center', vertical='center')
    # merge_cell.font = Font(name='Arial', size=12, bold=True, color='000000')
    merge_cell.fill = yellow_fill

    # 将工作簿保存为字节流
    stream = BytesIO()
    wb.save(stream)
    bytes = stream.getvalue()
    stream.close()
    return bytes


def get_excel_columns():
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
               'V', 'W', 'X', 'Y', 'Z']
    return columns


def generate_table(date, dpm=None):
    data = get_sales(date, dpm)
    rows = []
    if is_name_null(dpm):
        dpms, products = get_dpm_and_products_from_sales(data)
        price_count = 0
        row1 = [''] + [d for d in dpms] + ['数量', '价格', '总计']
        rows.append(row1)

        for p in products:
            row = [p]
            for i in range(1, len(row1) - 3):
                row.append(count_sale_with_dpm(p, row1[i], data))
            row.append(products[p]['num'])
            row.append(products[p]['price'])
            pc = products[p]['num'] * products[p]['price']
            row.append(pc)
            price_count += pc
            rows.append(row)

        row_end = ['总计']
        for i in range(1, len(row1) - 3):
            row_end.append(dpms[row1[i]])
        row_end.append('')
        row_end.append('')
        row_end.append(price_count)
        rows.append(row_end)

        row_old = ['养老金的券']
        row_income = ['收入']
        old_count = 0
        sale_with_old_pension = get_old_pensions(date)
        if sale_with_old_pension is None:
            sale_with_old_pension = []
        else:
            for s in sale_with_old_pension:
                pension = count_old_pension(s['comment'])
                if pension < 0:  # 如果计算出错，直接将最后一格填充：养老金的券格式有问题，无法计算
                    old_count = -1
                    sale_with_old_pension = []
                    break
                else:
                    old_count += pension

        for d in dpms:
            cell_value = ''
            for s in sale_with_old_pension:
                if s['dpm'] == d:
                    cell_value = s['comment']
                    break
            row_old.append(cell_value)
            row_income.append("")

        if old_count < 0:
            row_old += ["", "", "养老金的券格式有问题，无法计算"]
            row_income += ["", "", price_count]
        else:
            row_old += ["", "", old_count]
            row_income += ["", "", old_count + price_count]

        rows.append(row_old)
        rows.append(row_income)


    else:
        count = 0
        price_count = 0
        income = 0
        counter(0)

        apps, freeze = split_sales_by_type(data)
        row_app = [counter(), 'app', '下载', '备注', '价格', '总计']
        rows.append(row_app)
        for sale in apps:
            price = sale['num'] * sale['price']
            row = [counter(), sale['name'], sale['num'], sale['comment'], sale['price'], price]
            rows.append(row)
            count += sale['num']
            price_count += price
        row_app_end = [counter(), '总计', count, '', '', price_count]
        rows.append(row_app_end)
        income += price_count

        # clear count for app
        price_count = 0
        count = 0

        row_freeze = [counter(), '冻结', '下载', '备注', '价格', '总计']
        rows.append(row_freeze)
        for sale in freeze:
            price = sale['num'] * sale['price']
            row = [counter(), sale['name'], sale['num'], sale['comment'], sale['price'], price]
            rows.append(row)
            count += sale['num']
            price_count += price
        row_freeze_end = [counter(), '总计', count, '', '', price_count]
        rows.append(row_freeze_end)
        income += price_count

        row_end = [counter(), '收入', '', '', '', income]
        rows.append(row_end)
    return rows


def count_old_pension(comment):
    try:
        count = 0
        if comment is not None and comment.strip() != "":
            if "+" in comment:
                datas = comment.strip().split("+")
            else:
                datas = comment.strip()
            for data in datas:
                if "*" in comment:
                    cal = data.strip().split("*")
                    count += int(cal[0]) * int(cal[1])
        return count
    except Exception as e:
        return -1


def count_sale_with_dpm(pname, dpm, sales):
    count = 0
    for s in sales:
        if s['name'] == pname and s['dpm'] == dpm:
            count += s['num']
    if count == 0:
        return ''
    return count


def split_sales_by_type(sales):
    apps = []
    freeze = []
    for s in sales:
        if get_product_type(s['name']) == 'app':
            apps.append(s)
        else:
            freeze.append(s)
    return apps, freeze


def get_dpm_and_products_from_sales(sales):
    dpms = {}
    products = {}
    try:
        for s in sales:
            if s['dpm'] not in dpms:
                dpms[s['dpm']] = s['num']
            else:
                dpms[s['dpm']] += s['num']

            if s['name'] not in products:
                products[s['name']] = {'price': s['price'], 'num': s['num']}
            else:
                products[s['name']]['num'] += s['num']
    except Exception as e:
        log.error("error", e)
    return dpms, products


def get_department():
    table = TinyDB(dpm_db).table(dpm_table)
    return [item['name'] for item in table.all()]


def get_product():
    table = TinyDB(product_db).table(product_table)
    return table.all()


def add_departments(names):
    db = TinyDB(dpm_db)
    table = db.table(dpm_table)
    for name in names:
        table.insert({"name": name})
    db.close()


def add_products(products):
    db = TinyDB(product_db)
    table = db.table(product_table)
    for p in products:
        # table.insert({"name": p['name'], "price": p['price']})
        table.insert(p)
    db.close()


def del_department(name):
    db = TinyDB(dpm_db)
    table = db.table(dpm_table)
    Dpm = Query()
    table.remove(Dpm.name == name)
    db.close()


def del_product(name):
    db = TinyDB(product_db)
    table = db.table(product_table)
    Product = Query()
    table.remove(Product.name == name)
    db.close()


def update_product(p):
    db = TinyDB(product_db)
    table = db.table(product_table)
    Product = Query()
    table.update({"price": p['price'], "type": p['type']}, Product.name == p['name'])
    db.close()


def get_product_type(pname):
    db = TinyDB(product_db)
    table = db.table(product_table)
    Product = Query()
    p = table.get(Product.name == pname)
    return p['type']


def clear_sales(date, dpm=None):
    db = TinyDB(sale_db(date))
    table = db.table(sale_table)
    if is_name_null(dpm):
        table.truncate()
    else:
        Sale = Query()
        table.remove(Sale.dpm == dpm)
    db.close()


def add_sales(date, sales):
    db = TinyDB(sale_db(date))
    table = db.table(sale_table)
    table.insert_multiple(sales)
    db.close()


def get_sales(date, dpm=None):
    db = TinyDB(sale_db(date))
    table = db.table(sale_table)
    if is_name_null(dpm):
        data = table.all()
    else:
        Sale = Query()
        data = table.search(Sale.dpm == dpm)
    db.close()
    return data


def get_old_pensions(date):
    db = TinyDB(sale_db(date))
    table = db.table(sale_table)
    Sale = Query()
    data = table.search(Sale.comment != '')
    db.close()
    return data


def is_name_null(name):
    if name is None or name == '' or name == 'NA':
        return True
    return False


def sale_db(date):
    return '{}sale{}.json'.format(db_dir, date)


def counter(num=None):
    global row_count
    if num is None:
        row_count += 1
    else:
        row_count = num
    return row_count


if __name__ == '__main__':
    init()
    app.run(host='0.0.0.0', port=6060)
